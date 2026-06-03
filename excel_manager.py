"""
excel_manager.py
----------------
Gere o ficheiro Excel onde são guardadas as despesas.
Cria o ficheiro se não existir e adiciona despesas mensais.

Bibliotecas usadas:
- openpyxl : para criar e editar ficheiros Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os

# nome do ficheiro Excel
FICHEIRO = r"C:\Users\carlo\Desktop\despesas.xlsx"

# meses em português
MESES_PT = {
    "January": "Janeiro", "February": "Fevereiro", "March": "Março",
    "April": "Abril", "May": "Maio", "June": "Junho",
    "July": "Julho", "August": "Agosto", "September": "Setembro",
    "October": "Outubro", "November": "Novembro", "December": "Dezembro"
}

# categorias de despesas
DESPESAS_FIXAS = [
    "Renda",
    "Prestação do Carro",
    "Luz",
    "Água",
    "Gás",
    "Seguro de Saúde",
    "Seguro do Carro",
    "Telemóvel/Internet",
    "Escola/Creche",
    "Subscrições",
]

DESPESAS_VARIAVEIS = [
    "Compras",
    "Combustível",
    "Farmácia/Saúde",
    "Cabeleireiro/Higiene",
]


def mes_pt(data: datetime) -> str:
    """Devolve o nome do mês em português. Ex: 'Junho 2026'"""
    mes_en = data.strftime("%B")
    return f"{MESES_PT[mes_en]} {data.year}"


def obter_ou_criar_workbook():
    """
    Abre o ficheiro Excel se existir, ou cria um novo.
    """
    if os.path.exists(FICHEIRO):
        return openpyxl.load_workbook(FICHEIRO)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove a folha padrão
        return wb


def obter_ou_criar_folha(wb, mes_ano: str):
    """
    Obtém ou cria uma folha para o mês/ano indicado.
    Ex: "Junho 2026"
    """
    if mes_ano in wb.sheetnames:
        return wb[mes_ano]

    # cria folha nova com cabeçalhos
    ws = wb.create_sheet(mes_ano)

    # estilo do cabeçalho
    cor_cabecalho = PatternFill("solid", fgColor="2E86AB")
    fonte_cabecalho = Font(bold=True, color="FFFFFF", size=12)

    # cabeçalhos
    cabecalhos = ["Categoria", "Descrição", "Valor (€)", "Data"]
    for col, texto in enumerate(cabecalhos, 1):
        celula = ws.cell(row=1, column=col, value=texto)
        celula.fill = cor_cabecalho
        celula.font = fonte_cabecalho
        celula.alignment = Alignment(horizontal="center")

    # largura das colunas
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    return ws


def guardar_despesas(despesas: dict) -> bool:
    """
    Guarda as despesas no Excel.
    despesas = { "Renda": 500.0, "Luz": 45.0, ... }
    """
    try:
        wb = obter_ou_criar_workbook()

        # nome do mês atual em português
        agora = datetime.now()
        mes_ano = mes_pt(agora)
        data_hoje = agora.strftime("%d/%m/%Y")

        ws = obter_ou_criar_folha(wb, mes_ano)

        # encontra a próxima linha vazia
        proxima_linha = ws.max_row + 1

        # cor alternada para as linhas
        cor_fixa = PatternFill("solid", fgColor="D4E6F1")
        cor_variavel = PatternFill("solid", fgColor="D5F5E3")

        for categoria, valor in despesas.items():
            if valor is None or valor == 0:
                continue

            # determina se é fixa ou variável
            cor = cor_fixa if categoria in DESPESAS_FIXAS else cor_variavel

            # escreve a linha
            linha = [categoria, "", valor, data_hoje]
            for col, dado in enumerate(linha, 1):
                celula = ws.cell(row=proxima_linha, column=col, value=dado)
                celula.fill = cor
                if col == 3:  # coluna do valor
                    celula.number_format = "#,##0.00 €"

            proxima_linha += 1

        # linha de total
        proxima_linha += 1
        total = sum(v for k, v in despesas.items() if isinstance(v, (int, float)))

        ws.cell(row=proxima_linha, column=2, value="TOTAL DO MÊS").font = Font(bold=True)
        celula_total = ws.cell(row=proxima_linha, column=3, value=total)
        celula_total.font = Font(bold=True, color="E74C3C")
        celula_total.number_format = "#,##0.00 €"

        wb.save(FICHEIRO)
        print(f"✅ Despesas guardadas em {FICHEIRO}")
        return True

    except Exception as erro:
        print(f"❌ Erro ao guardar: {erro}")
        return False


def obter_resumo_mes_anterior() -> dict:
    """
    Lê o Excel e devolve o resumo do mês anterior.
    Devolve: { "mes": "Maio 2026", "despesas": { "Renda": 500.0, ... }, "total": 1234.56 }
    """
    try:
        if not os.path.exists(FICHEIRO):
            return {}

        # calcula o mês anterior
        mes_anterior = datetime.now() - relativedelta(months=1)
        nome_mes = mes_pt(mes_anterior)

        wb = openpyxl.load_workbook(FICHEIRO)

        if nome_mes not in wb.sheetnames:
            return {"mes": nome_mes, "despesas": {}, "total": 0.0}

        ws = wb[nome_mes]
        despesas = {}
        total = 0.0

        for row in ws.iter_rows(min_row=2, values_only=True):
            categoria = row[0]
            valor = row[2]

            if categoria and valor and isinstance(valor, (int, float)):
                # agrupa por categoria somando valores
                if categoria in despesas:
                    despesas[categoria] += valor
                else:
                    despesas[categoria] = valor
                total += valor

        return {
            "mes": nome_mes,
            "despesas": despesas,
            "total": total
        }

    except Exception as erro:
        print(f"❌ Erro ao ler resumo: {erro}")
        return {}


def obter_total_mes(mes_ano: str = None) -> float:
    """
    Obtém o total de despesas de um mês.
    Se não for indicado, usa o mês atual.
    """
    try:
        if not os.path.exists(FICHEIRO):
            return 0.0

        if mes_ano is None:
            mes_ano = mes_pt(datetime.now())

        wb = openpyxl.load_workbook(FICHEIRO)

        if mes_ano not in wb.sheetnames:
            return 0.0

        ws = wb[mes_ano]
        total = 0.0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] and isinstance(row[2], (int, float)):
                total += row[2]

        return total

    except Exception as erro:
        print(f"❌ Erro: {erro}")
        return 0.0